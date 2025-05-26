import os
import json
import pandas as pd
import argparse
import numpy as np
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Alignment

def calculate_entropy(probabilities):
    """
    Calculate the entropy of a probability distribution.
    
    :param probabilities: list of float, probability values
    :return: float, entropy value
    """
    probabilities = np.array(probabilities)
    probabilities = probabilities[probabilities > 0]  # Exclude zero probabilities to avoid log2(0)
    return -np.sum(probabilities * np.log2(probabilities))

def calculate_accuracy(predictions, true_labels):
    """
    Calculate accuracy based on predictions and true labels.
    
    :param predictions: list of str, predicted labels
    :param true_labels: list of str, true labels
    :return: float, accuracy percentage
    """
    if len(predictions) != len(true_labels):
        raise ValueError("Number of predictions and true labels must be the same")

    correct = sum(1 for pred, true in zip(predictions, true_labels) if pred == true)
    total = len(predictions)
    accuracy = correct / total
    scaled_accuracy = accuracy * 100
    return round(scaled_accuracy, 2)

def calculate_diffTM(predictions, true_labels, replaced_labels):
    """
    Calculate extent of LLMs’ tendency to rely on misinformation polluted contexts over correct contexts
    
    :param predictions: list of str, predicted labels
    :param true_labels: list of str, true labels
    :param replaced_labels: list of str, replaced labels
    :return: 
    """
    if len(predictions) != len(true_labels) and len(predictions) != len(replaced_labels):
        raise ValueError("Number of predictions and labels must be the same")
    
    total = len(predictions)
    correct = sum(1 for pred, true in zip(predictions, true_labels) if pred == true)
    false = sum(1 for pred, replaced in zip(predictions, replaced_labels) if pred == replaced)
    correct_ratio = correct / total
    false_ratio = false / total
    diffTM = (false_ratio - correct_ratio) / (false_ratio + correct_ratio)  # range: [-1, 1]
    return round(diffTM, 2)

def process_json_file_with_origin(json_file_path, default_root):
    """
    Process a single JSON file to extract predictions and calculate metrics.
    
    :param json_file_path: str, path to the JSON file
    :param default_root: str, root directory containing baseline JSON files
    :return: tuple of metrics (correct_accuracy, replaced_accuracy, uncertain_accuracy, average_entropy)
    """
    selected_indices = []
    cnt = 0

    baseline1 = os.path.join(default_root, 'default_style_obj_choice.jsonl')

    selected_indices1 = []
    with open(baseline1, 'r', encoding='utf-8') as f:
        for line in f:
            data = json.loads(line)
            if data['prediction'] == data['correct_option']:
                selected_indices1.append(cnt)
            cnt += 1

    selected_indices = list(set(selected_indices1))          
            
    predictions = []
    true_labels = []
    replaced_labels = []
    irrelavant_labels = []
    uncertain_labels = []
    none_labels = []
    entropies = []
    cnt = 0

    print(f"Processing {default_root}!")
    print(f"Default correct predictions: {len(selected_indices1)}!")
    print(f"Total test cases: {len(selected_indices)}!")

    with open(json_file_path, 'r') as f:
        for line in f:
            data = json.loads(line)
            if cnt in selected_indices:
                predictions.append(data['prediction'])
                true_labels.append(data['correct_option'])
                replaced_labels.append(data['replaced_option'])
                irrelavant_labels.append(data['irrelavant_option'])
                uncertain_labels.append(data['not_sure_option'])
                none_labels.append(data['not_in_option_option'])
                probabilities = list(data["prob"].values())
                entropy = calculate_entropy(probabilities)
                entropies.append(entropy)
            
            cnt += 1

    correct_accuracy = calculate_accuracy(predictions, true_labels)
    replaced_accuracy = calculate_accuracy(predictions, replaced_labels)
    irrelavant_accuracy = calculate_accuracy(predictions, irrelavant_labels)
    uncertain_accuracy = calculate_accuracy(predictions, uncertain_labels)
    none_accuracy = calculate_accuracy(predictions, none_labels)
    average_entropy = round(np.mean(entropies), 2)
    diffTM = calculate_diffTM(predictions, true_labels, replaced_labels)
    
    return correct_accuracy, replaced_accuracy, irrelavant_accuracy, uncertain_accuracy, none_accuracy, average_entropy, diffTM
    

def process_json_file_with_origin_TF(json_file_path, default_root):
    """
    Process a single JSON file to extract predictions and calculate metrics.

    filter out those correct predictions in the default files and those false predictions in the default files
    
    :param json_file_path: str, path to the JSON file
    :param default_root: str, root directory containing baseline JSON files
    :return: tuple of metrics (correct_accuracy, replaced_accuracy, uncertain_accuracy, average_entropy)
    """
    selected_indices = []
    cnt = 0

    baseline1 = os.path.join(default_root, 'default_style_obj_choice.jsonl')

    selected_indices1 = []
    with open(baseline1, 'r', encoding='utf-8') as f:
        for line in f:
            data = json.loads(line)
            if data['prediction'] == data['correct_option']:
                selected_indices1.append(cnt)
            cnt += 1


    selected_indices = list(set(selected_indices1))        

    print(f"Processing {json_file_path}!")
    print(f"Default correct predictions: {len(selected_indices)}!")
    print(f"Default false predictions: {cnt - len(selected_indices)}!")  
    
    # For those correct predictions in the default files
    predictions = []
    true_labels = []
    replaced_labels = []
    irrelavant_labels = []
    uncertain_labels = []
    none_labels = []
    entropies = []

    # For those false predictions in the default files
    predictions_F = []
    true_labels_F = []
    replaced_labels_F = []
    irrelavant_labels_F = []
    uncertain_labels_F = []
    none_labels_F = []
    entropies_F = []

    cnt = 0

    with open(json_file_path, 'r') as f:
        total_lines = sum(1 for line in f)

    with open(json_file_path, 'r') as f:
        for line in tqdm(f, total=total_lines, desc="Processing"):
            data = json.loads(line)
            if cnt in selected_indices:
                predictions.append(data['prediction'])
                true_labels.append(data['correct_option'])
                replaced_labels.append(data['replaced_option'])
                irrelavant_labels.append(data['irrelavant_option'])
                uncertain_labels.append(data['not_sure_option'])
                none_labels.append(data['not_in_option_option'])
                probabilities = list(data["prob"].values())
                entropy = calculate_entropy(probabilities)
                entropies.append(entropy)
            else:
                predictions_F.append(data['prediction'])
                true_labels_F.append(data['correct_option'])
                replaced_labels_F.append(data['replaced_option'])
                irrelavant_labels_F.append(data['irrelavant_option'])
                uncertain_labels_F.append(data['not_sure_option'])
                none_labels_F.append(data['not_in_option_option'])
                probabilities = list(data["prob"].values())
                entropy = calculate_entropy(probabilities)
                entropies_F.append(entropy)
            
            cnt += 1

    correct_accuracy = calculate_accuracy(predictions, true_labels)
    replaced_accuracy = calculate_accuracy(predictions, replaced_labels)
    irrelavant_accuracy = calculate_accuracy(predictions, irrelavant_labels)
    uncertain_accuracy = calculate_accuracy(predictions, uncertain_labels)
    none_accuracy = calculate_accuracy(predictions, none_labels)
    average_entropy = round(np.mean(entropies), 2)
    diffTM = calculate_diffTM(predictions, true_labels, replaced_labels)
    metrics_with_correct_origin = {
        'Correct Accuracy': correct_accuracy,
        'Replaced Accuracy': replaced_accuracy,
        'Irrelavant Accuracy': irrelavant_accuracy,
        'Uncertain Accuracy': uncertain_accuracy,
        'None Accuracy': none_accuracy,
        'DiffTM': diffTM,
        'Average Entropy': average_entropy
    }

    correct_accuracy_F = calculate_accuracy(predictions_F, true_labels_F)
    replaced_accuracy_F = calculate_accuracy(predictions_F, replaced_labels_F)
    irrelavant_accuracy_F = calculate_accuracy(predictions_F, irrelavant_labels_F)
    uncertain_accuracy_F = calculate_accuracy(predictions_F, uncertain_labels_F)
    none_accuracy_F = calculate_accuracy(predictions_F, none_labels_F)
    average_entropy_F = round(np.mean(entropies_F), 2)
    diffTM_F = calculate_diffTM(predictions_F, true_labels_F, replaced_labels_F)
    metrics_with_false_origin = {
        'Correct Accuracy': correct_accuracy_F,
        'Replaced Accuracy': replaced_accuracy_F,
        'Irrelavant Accuracy': irrelavant_accuracy_F,
        'Uncertain Accuracy': uncertain_accuracy_F,
        'None Accuracy': none_accuracy_F,
        'DiffTM': diffTM_F,
        'Average Entropy': average_entropy_F
    }
    
    return metrics_with_correct_origin, metrics_with_false_origin


def process_json_file(json_file_path, default_root):
    """
    Process a single JSON file to extract predictions and calculate metrics.
    [We don not need to filter those correct predictions in the default and correct files]
    
    :param json_file_path: str, path to the JSON file
    :param default_root: str, root directory containing baseline JSON files
    :return: tuple of metrics (correct_accuracy, replaced_accuracy, uncertain_accuracy, average_entropy)
    """

    predictions = []
    true_labels = []
    replaced_labels = []
    irrelavant_labels = []
    uncertain_labels = []
    none_labels = []
    entropies = []

    print(f"Processing {default_root}!")

    with open(json_file_path, 'r') as f:
        for line in f:
            data = json.loads(line)
            predictions.append(data['prediction'])
            true_labels.append(data['correct_option'])
            replaced_labels.append(data['replaced_option'])
            irrelavant_labels.append(data['irrelavant_option'])
            uncertain_labels.append(data['not_sure_option'])
            none_labels.append(data['not_in_option_option'])
            probabilities = list(data["prob"].values())
            entropy = calculate_entropy(probabilities)
            entropies.append(entropy)

    correct_accuracy = calculate_accuracy(predictions, true_labels)
    replaced_accuracy = calculate_accuracy(predictions, replaced_labels)
    irrelavant_accuracy = calculate_accuracy(predictions, irrelavant_labels)
    uncertain_accuracy = calculate_accuracy(predictions, uncertain_labels)
    none_accuracy = calculate_accuracy(predictions, none_labels)
    average_entropy = round(np.mean(entropies), 2)
    diffTM = calculate_diffTM(predictions, true_labels, replaced_labels)
    
    return correct_accuracy, replaced_accuracy, irrelavant_accuracy, uncertain_accuracy, none_accuracy, average_entropy, diffTM

def save_to_excel_MR(results, out_file):
    """
    Save the evaluation results to an Excel file.
    [Memorization Ratio Settting]
    
    :param results: dict, evaluation results
    :param out_file: str, path to the output Excel file
    """
    wb = Workbook()
    ws = wb.active

    headers = []
    sub_headers = []
    for category in sorted(results.keys()):
        headers.extend([category, '', '', '', ''])
        sub_headers.extend(['Correct Accuracy', 'Replaced Accuracy', 'Uncertain Accuracy', 'Memorization Ratio', 'Average Entropy'])
    
    ws.append(headers)
    ws.append(sub_headers)
    
    col_idx = 1
    for category in sorted(results.keys()):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 4)
        col_idx += 5
    
    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    max_len = max(len(v) for v in results.values())
    for i in range(max_len):
        row = []
        for category in sorted(results.keys()):
            if i < len(results[category]):
                row.extend(results[category][i].values())
            else:
                row.extend([''] * 5)
        ws.append(row)

    os.makedirs(os.path.dirname(out_file), exist_ok=True)
    wb.save(out_file)

def save_to_excel_EM(results, out_file):
    """
    Save the evaluation results to an Excel file.
    [Exactly Match Setting in Question-Choice]
    
    :param results: dict, evaluation results
    :param out_file: str, path to the output Excel file
    """
    wb = Workbook()
    ws = wb.active

    headers = []
    sub_headers = []
    for category in sorted(results.keys()):
        headers.extend([category, '', '', '', '', '', ''])
        sub_headers.extend(['Correct Accuracy', 'Replaced Accuracy', 'Irrelavant Accuracy', 'Uncertain Accuracy', 'None Accuracy', 'DiffTM', 'Average Entropy'])
    
    ws.append(headers)
    ws.append(sub_headers)
    
    col_idx = 1
    for category in sorted(results.keys()):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 6)
        col_idx += 7
    
    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    max_len = max(len(v) for v in results.values())
    for i in range(max_len):
        row = []
        for category in sorted(results.keys()):
            if i < len(results[category]):
                row.extend(results[category][i].values())
            else:
                row.extend([''] * 7)
        ws.append(row)

    os.makedirs(os.path.dirname(out_file), exist_ok=True)
    wb.save(out_file)


def process_directory_EM_with_origin_TF(directory, out_dir):
    """
    Process a directory of JSON files and evaluate the results.
    fiter those correct/false predictions in the default files
    [Exactly Match in Question-Choice]
    
    :param directory: str, path to the directory containing JSON files
    :param out_dir: str, path to the output directory
    """
    default_root = directory

    # for those correct predictions in the default files
    default_results = {}
    context_conflict_results = {}
    inter_conflict_results = {}
    description_results = {}

    # for those false predictions in the default files
    default_results_F = {}
    context_conflict_results_F = {}
    inter_conflict_results_F = {}
    description_results_F = {}
    
    for root, dirs, files in os.walk(directory):
        model_name = os.path.basename(root)
        for file in files:
            if file.endswith(".jsonl"):
                file_path = os.path.join(root, file)
                file_name = os.path.splitext(file)[0]
                category = file_name.split('_style')[0]
                style = file_name.split('style_')[1].split('_choice')[0]
                if category not in ['fact']:
                    continue
            

                result_entry_correct, result_entry_false = process_json_file_with_origin_TF(file_path, default_root)

                if style in ['obj', 'blog', 'news_report', 'science_reference', 'social_media', 'confident_language', 'technical_language']:
                    if model_name not in context_conflict_results:
                        context_conflict_results[model_name] = {}
                    if category not in context_conflict_results[model_name]:
                        context_conflict_results[model_name][f'{category}_{style}'] = []
                    context_conflict_results[model_name][f'{category}_{style}'].append(result_entry_correct)
                
                if style in ['obj', 'blog', 'news_report', 'science_reference', 'social_media', 'confident_language', 'technical_language']:
                    if model_name not in context_conflict_results_F:
                        context_conflict_results_F[model_name] = {}
                    if category not in context_conflict_results_F[model_name]:
                        context_conflict_results_F[model_name][f'{category}_{style}'] = []
                    context_conflict_results_F[model_name][f'{category}_{style}'].append(result_entry_false)
                
    
    output_path = os.path.join(out_dir, model_name)
    output_path_T = os.path.join(output_path, 'with_correct_origin')
    output_path_F = os.path.join(output_path, 'with_false_origin')
    os.makedirs(output_path, exist_ok=True)



    for model_name, results in context_conflict_results.items():
        out_file = os.path.join(output_path_T, 'context_conflict_style.xlsx')
        save_to_excel_EM(results, out_file)
    


    
    for model_name, results in context_conflict_results_F.items():
        out_file = os.path.join(output_path_F, 'context_conflict_style.xlsx')
        save_to_excel_EM(results, out_file)
    


def process_directory_EM_with_origin(directory, out_dir):
    """
    Process a directory of JSON files and evaluate the results.
    fiter those correct predictions in the default files
    [Exactly Match in Question-Choice]
    
    :param directory: str, path to the directory containing JSON files
    :param out_dir: str, path to the output directory
    """
    default_root = directory
    default_results = {}
    context_conflict_results = {}
    inter_conflict_results = {}
    description_results = {}
    
    for root, dirs, files in os.walk(directory):
        model_name = os.path.basename(root)
        for file in files:
            if file.endswith(".jsonl"):
                file_path = os.path.join(root, file)
                file_name = os.path.splitext(file)[0]
                category = file_name.split('_style')[0]
                
                correct_accuracy, replaced_accuracy, irrelavant_accuracy, uncertain_accuracy, none_accuracy, average_entropy, diffTM = process_json_file_with_origin(file_path, default_root)

                
                result_entry = {
                    'Correct Accuracy': correct_accuracy,
                    'Replaced Accuracy': replaced_accuracy,
                    'Irrelavant Accuracy': irrelavant_accuracy,
                    'Uncertain Accuracy': uncertain_accuracy,
                    'None Accuracy': none_accuracy,
                    'DiffTM': diffTM,
                    'Average Entropy': average_entropy
                }

                if category in ['default', 'correct']:
                    if model_name not in default_results:
                        default_results[model_name] = {}
                    if category not in default_results[model_name]:
                        default_results[model_name][category] = []
                    default_results[model_name][category].append(result_entry)

                # 内外部知识冲突
                if category in ['fact', 'temporal', 'semantic']:
                    if model_name not in context_conflict_results:
                        context_conflict_results[model_name] = {}
                    if category not in context_conflict_results[model_name]:
                        context_conflict_results[model_name][category] = []
                    context_conflict_results[model_name][category].append(result_entry)
                
                # 外部多个知识冲突
                if category in ['correct_mis', 'correct_temporal', 'correct_semantic']:
                    if model_name not in inter_conflict_results:
                        inter_conflict_results[model_name] = {}
                    if category not in inter_conflict_results[model_name]:
                        inter_conflict_results[model_name][category] = []
                    inter_conflict_results[model_name][category].append(result_entry)
                
    
    output_path = os.path.join(out_dir, model_name)
    output_path = os.path.join(output_path, 'with_origin')
    os.makedirs(output_path, exist_ok=True)

    for model_name, results in default_results.items():
        out_file = os.path.join(output_path, 'default.xlsx')
        save_to_excel_EM(results, out_file)

    for model_name, results in context_conflict_results.items():
        out_file = os.path.join(output_path, 'context_conflict.xlsx')
        save_to_excel_EM(results, out_file)

    for model_name, results in inter_conflict_results.items():
        out_file = os.path.join(output_path, 'inter_conflict.xlsx')
        save_to_excel_EM(results, out_file)

    for model_name, results in description_results.items():
        out_file = os.path.join(output_path, 'description.xlsx')
        save_to_excel_EM(results, out_file)


def process_directory_EM(directory, out_dir):
    """
    Process a directory of JSON files and evaluate the results.
    [Exactly Match in Question-Choice]
    
    :param directory: str, path to the directory containing JSON files
    :param out_dir: str, path to the output directory
    """
    default_root = directory
    default_results = {}
    context_conflict_results = {}
    inter_conflict_results = {}
    description_results = {}
    
    for root, dirs, files in os.walk(directory):
        model_name = os.path.basename(root)
        for file in files:
            if file.endswith(".jsonl"):
                file_path = os.path.join(root, file)
                file_name = os.path.splitext(file)[0]
                category = file_name.split('_style')[0]
                
                correct_accuracy, replaced_accuracy, irrelavant_accuracy, uncertain_accuracy, none_accuracy, average_entropy, diffTM = process_json_file(file_path, default_root)

                
                result_entry = {
                    'Correct Accuracy': correct_accuracy,
                    'Replaced Accuracy': replaced_accuracy,
                    'Irrelavant Accuracy': irrelavant_accuracy,
                    'Uncertain Accuracy': uncertain_accuracy,
                    'None Accuracy': none_accuracy,
                    'DiffTM': diffTM,
                    'Average Entropy': average_entropy
                }

                # TODO: modify to adopt my files
                if category in ['default', 'correct']:
                    if model_name not in default_results:
                        default_results[model_name] = {}
                    if category not in default_results[model_name]:
                        default_results[model_name][category] = []
                    default_results[model_name][category].append(result_entry)

                # 内外部知识冲突
                if category in ['fact', 'temporal', 'semantic']:
                    if model_name not in context_conflict_results:
                        context_conflict_results[model_name] = {}
                    if category not in context_conflict_results[model_name]:
                        context_conflict_results[model_name][category] = []
                    context_conflict_results[model_name][category].append(result_entry)
                
                # 外部多个知识冲突
                if category in ['correct_mis', 'correct_temporal', 'correct_semantic']:
                    if model_name not in inter_conflict_results:
                        inter_conflict_results[model_name] = {}
                    if category not in inter_conflict_results[model_name]:
                        inter_conflict_results[model_name][category] = []
                    inter_conflict_results[model_name][category].append(result_entry)
    

    output_path = os.path.join(out_dir, model_name)
    os.makedirs(output_path, exist_ok=True)

    for model_name, results in default_results.items():
        out_file = os.path.join(output_path, 'default.xlsx')
        save_to_excel_EM(results, out_file)

    for model_name, results in context_conflict_results.items():
        out_file = os.path.join(output_path, 'context_conflict.xlsx')
        save_to_excel_EM(results, out_file)

    for model_name, results in inter_conflict_results.items():
        out_file = os.path.join(output_path, 'inter_conflict.xlsx')
        save_to_excel_EM(results, out_file)

    for model_name, results in description_results.items():
        out_file = os.path.join(output_path, 'description.xlsx')
        save_to_excel_EM(results, out_file)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Evaluate model performance on JSON datasets and save results to Excel.")
    parser.add_argument("--input_dir", type=str, help="Directory containing the JSON files (inference results).")
    parser.add_argument("--out_dir", type=str, help="Directory to save the output Excel files.")
    args = parser.parse_args()
    # process_directory_EM(args.input_dir, args.out_dir)
    # process_directory_EM_with_origin(args.input_dir, args.out_dir)
    process_directory_EM_with_origin_TF(args.input_dir, args.out_dir)
